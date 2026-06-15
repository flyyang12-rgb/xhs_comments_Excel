import time
import json
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile
from DrissionPage import WebPage
from openpyxl import Workbook, load_workbook

# 关键词
keywords = [
    "儿童运动内裤",
    "男童运动内裤",
    "女童运动内裤",
    "儿童速干内裤",
    "儿童排汗内裤",
    "儿童一片式内裤",
    "儿童透气运动内裤",
    "儿童运动内裤推荐"
]


def search_keyword(tab, keyword):
    url = f'https://www.xiaohongshu.com/search_result_ai?keyword={keyword}&source=web_explore_feed&type=51'

    tab.get(url)
    time.sleep(3)

    # 悬浮筛选
    tab.ele('text=筛选').hover()

    time.sleep(1)

    # 点击最多评论
    tab.ele('text=最多评论').click()

    time.sleep(5)

    # 开启监听
    tab.listen.start(['api/sns/web/v2/search/notes'])

    # 再点一次最多评论
    tab.ele('text=最多评论').click()

    response = tab.listen.wait(timeout=15)

    json_data = response.response.body

    return json_data


def parse_notes(json_data, keyword):
    if isinstance(json_data, dict):
        data = json_data
    else:
        data = json.loads(json_data)

    items = data['data']['items']

    result = []

    batch = datetime.now().strftime(f'%Y%m%d_{keyword}_01')
    collect_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    rank = 0

    for item in items:

        if item.get('model_type') != 'note':
            continue

        note = item.get('note_card', {})

        note_id = item.get('id')
        xsec_token = item.get('xsec_token')

        if not note_id or not xsec_token:
            continue

        rank += 1

        # 只取前10条
        if rank > 10:
            break

        link = f"https://www.xiaohongshu.com/explore/{note_id}?xsec_token={xsec_token}"

        title = note.get('display_title', '无标题')

        author = note.get('user', {}).get('nick_name', '未知作者')

        comment_count = note.get('interact_info', {}).get('comment_count', '0')

        row = [
            batch,
            collect_time,
            keyword,
            rank,
            link,
            title,
            author,
            comment_count,
            '成功'
        ]

        result.append(row)

    return result


def save_to_excel(all_rows):
    output_file = Path('notes_raw.xlsx')
    headers = [
        '采集批次',
        '采集时间',
        '搜索关键词',
        '关键词下排名',
        '笔记链接',
        '笔记标题',
        '作者昵称',
        '评论数',
        '采集状态'
    ]

    try:
        wb = load_workbook(output_file)
    except (FileNotFoundError, BadZipFile):
        wb = Workbook()
        wb.active.title = 'notes_raw'

    if 'notes_raw' in wb.sheetnames:
        ws = wb['notes_raw']
    else:
        ws = wb.active
        ws.title = 'notes_raw'

    first_row = [cell.value for cell in ws[1]]
    if first_row != headers:
        if ws.max_row == 1 and all(value is None for value in first_row):
            for col_index, value in enumerate(headers, start=1):
                ws.cell(row=1, column=col_index, value=value)
        else:
            ws.insert_rows(1)
            for col_index, value in enumerate(headers, start=1):
                ws.cell(row=1, column=col_index, value=value)

    for row in all_rows:
        ws.append(row)

    wb.save(output_file)


def main():
    wp = WebPage()

    tab = wp.latest_tab

    all_rows = []

    for keyword in keywords:

        print(f'开始采集：{keyword}')

        try:
            json_data = search_keyword(tab, keyword)

            rows = parse_notes(json_data, keyword)

            all_rows.extend(rows)

            print(f'{keyword} 采集完成，共 {len(rows)} 条')

        except Exception as e:
            print(f'{keyword} 采集失败：{e}')

    save_to_excel(all_rows)

    print('全部采集完成')


if __name__ == '__main__':
    main()
