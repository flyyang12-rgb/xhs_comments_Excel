import argparse
import json
import random
import re
import shutil
import subprocess
import time
import urllib.parse
from pathlib import Path

import requests
from DrissionPage import WebPage
from openpyxl import Workbook, load_workbook


COMMENTS_RAW_HEADERS = [
    '笔记链接',
    '评论序号',
    '一级评论内容和图片链接',
    '所有二级评论内容和图片链接',
    '评论采集状态',
]
DONE_MARK = '评论采集完成'
BASE_URL = 'https://edith.xiaohongshu.com'
ROOT = Path(__file__).resolve().parent


def parse_note_url(note_url):
    parsed = urllib.parse.urlparse(note_url)
    note_id = parsed.path.rstrip('/').split('/')[-1]
    params = urllib.parse.parse_qs(parsed.query)
    xsec_token = params.get('xsec_token', [''])[0]
    if not re.fullmatch(r'[0-9a-fA-F]{24}', note_id or ''):
        raise ValueError('无效笔记链接')
    return note_id, xsec_token


def trans_cookies(cookies_str):
    cookies = {}
    for item in cookies_str.split(';'):
        item = item.strip()
        if item and '=' in item:
            key, value = item.split('=', 1)
            cookies[key] = value
    return cookies


def cookies_to_str(cookies):
    if isinstance(cookies, str):
        return cookies
    if isinstance(cookies, dict):
        return '; '.join(f'{key}={value}' for key, value in cookies.items())
    return '; '.join(
        f"{cookie['name']}={cookie.get('value', '')}"
        for cookie in (cookies or [])
        if isinstance(cookie, dict) and cookie.get('name')
    )


def get_browser_cookies(note_url):
    page = WebPage()
    tab = page.latest_tab
    tab.get(note_url)
    time.sleep(10)
    cookies = None
    for obj in (tab, page):
        getter = getattr(obj, 'cookies', None)
        if not getter:
            continue
        try:
            cookies = getter(as_dict=False)
        except TypeError:
            cookies = getter()
        if cookies:
            break
    cookies_str = cookies_to_str(cookies)
    if 'a1=' not in cookies_str:
        input('请在打开的浏览器里登录小红书，登录完成后回到这里按回车继续...')
        try:
            cookies_str = cookies_to_str(tab.cookies(as_dict=False))
        except TypeError:
            cookies_str = cookies_to_str(tab.cookies())
    return cookies_str


def sign_headers(api, cookies_str):
    cookies = trans_cookies(cookies_str)
    a1 = cookies.get('a1')
    if not a1:
        raise RuntimeError('缺少 a1 cookie')
    payload = {'api': api, 'data': '', 'a1': a1, 'method': 'GET'}
    completed = subprocess.run(
        ['node', str(ROOT / 'xhs_sign.js')],
        input=json.dumps(payload, ensure_ascii=False),
        text=True,
        capture_output=True,
        check=True,
        cwd=str(ROOT),
    )
    signed = json.loads(completed.stdout)
    return {
        'authority': 'edith.xiaohongshu.com',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'zh-CN,zh;q=0.9',
        'origin': 'https://www.xiaohongshu.com',
        'referer': 'https://www.xiaohongshu.com/',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'x-b3-traceid': ''.join(random.choice('abcdef0123456789') for _ in range(16)),
        'x-s': signed['xs'],
        'x-s-common': signed['xs_common'],
        'x-t': str(signed['xt']),
    }


def request_get(api, params, cookies_str):
    splice_api = f'{api}?{urllib.parse.urlencode(params)}'
    headers = sign_headers(splice_api, cookies_str)
    session = requests.Session()
    session.trust_env = False
    response = session.get(
        BASE_URL + splice_api,
        headers=headers,
        cookies=trans_cookies(cookies_str),
        timeout=20,
    )
    response.raise_for_status()
    data = response.json()
    if not data.get('success'):
        raise RuntimeError(data.get('msg') or '接口返回失败')
    return data


def get_out_comments(note_id, xsec_token, cookies_str, limit=0):
    cursor = ''
    comments = []
    while True:
        data = request_get('/api/sns/web/v2/comment/page', {
            'note_id': note_id,
            'cursor': cursor,
            'top_comment_id': '',
            'image_formats': 'jpg,webp,avif',
            'xsec_token': xsec_token,
        }, cookies_str)
        page_comments = data.get('data', {}).get('comments') or []
        if limit and limit > 0:
            remaining = limit - len(comments)
            comments.extend(page_comments[:remaining])
        else:
            comments.extend(page_comments)
        print(f'已获取一级评论 {len(comments)} 条', flush=True)
        if (limit and limit > 0 and len(comments) >= limit) or not data.get('data', {}).get('has_more') or not page_comments:
            break
        cursor = str(data.get('data', {}).get('cursor') or '')
        time.sleep(15)
    return comments


def get_sub_comments(comment, xsec_token, cookies_str):
    sub_comments = list(comment.get('sub_comments') or [])
    seen_ids = {
        sub_comment.get('id')
        for sub_comment in sub_comments
        if isinstance(sub_comment, dict) and sub_comment.get('id')
    }

    if not comment.get('sub_comment_has_more'):
        return sub_comments

    cursor = comment.get('sub_comment_cursor') or ''
    while True:
        data = request_get('/api/sns/web/v2/comment/sub/page', {
            'note_id': comment.get('note_id', ''),
            'root_comment_id': comment.get('id', ''),
            'num': '10',
            'cursor': cursor,
            'image_formats': 'jpg,webp,avif',
            'top_comment_id': '',
            'xsec_token': xsec_token,
        }, cookies_str)
        page_comments = data.get('data', {}).get('comments') or []
        for sub_comment in page_comments:
            sub_id = sub_comment.get('id')
            if sub_id and sub_id in seen_ids:
                continue
            if sub_id:
                seen_ids.add(sub_id)
            sub_comments.append(sub_comment)

        if not data.get('data', {}).get('has_more') or not page_comments:
            break
        cursor = str(data.get('data', {}).get('cursor') or '')
        time.sleep(15)

    return sub_comments


def get_all_comments_with_replies(note_id, xsec_token, cookies_str, limit=0):
    comments = get_out_comments(note_id, xsec_token, cookies_str, limit=limit)
    for index, comment in enumerate(comments, start=1):
        comment['sub_comments'] = get_sub_comments(comment, xsec_token, cookies_str)
        if index % 20 == 0 or index == len(comments):
            print(f'已处理回复 {index}/{len(comments)}', flush=True)
    return comments


def header_map(ws):
    return {
        str(ws.cell(1, col).value or '').strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value is not None
    }


def cell(ws, row, headers, name):
    col = headers.get(name)
    return ws.cell(row, col).value if col else None


def first_valid_note(ws, headers):
    for row in range(2, ws.max_row + 1):
        status = str(cell(ws, row, headers, '采集状态') or '').strip()
        link = str(cell(ws, row, headers, '笔记链接') or '').strip()
        if status == '成功' and link:
            return link
    return ''


def extract_picture_links(comment):
    links = []
    for picture in comment.get('pictures') or []:
        if not isinstance(picture, dict):
            continue
        url = picture.get('url_default') or picture.get('url_pre')
        if not url:
            for item in picture.get('info_list') or []:
                if isinstance(item, dict) and item.get('url'):
                    url = item['url']
                    break
        if url:
            links.append(url)
    return '\n'.join(links)


def build_single_comment_field(comment):
    lines = []
    content = (comment.get('content') or '').strip()
    if content:
        lines.append(content)
    picture_links = extract_picture_links(comment)
    if picture_links:
        lines.append(picture_links)

    return '\n'.join(lines)


def build_replies_field(comment):
    blocks = []
    for index, sub_comment in enumerate(comment.get('sub_comments') or [], start=1):
        field = build_single_comment_field(sub_comment)
        if field:
            blocks.append(f'回复{index}: {field}')
    return '\n'.join(blocks)


def comments_raw_sheet(output_path):
    if output_path.exists():
        wb = load_workbook(output_path)
        ws = wb['comments_raw'] if 'comments_raw' in wb.sheetnames else wb.active
        existing_headers = [ws.cell(1, col).value for col in range(1, len(COMMENTS_RAW_HEADERS) + 1)]
        if existing_headers != COMMENTS_RAW_HEADERS:
            ws = wb.create_sheet('comments_raw')
            ws.append(COMMENTS_RAW_HEADERS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'comments_raw'
        ws.append(COMMENTS_RAW_HEADERS)
    return wb, ws


def ensure_progress_column(ws):
    if ws.cell(1, 10).value != DONE_MARK:
        ws.cell(1, 10).value = DONE_MARK
    return 10


def append_status_row(out_ws, note, note_id, status):
    out_ws.append([
        note.get('笔记链接', ''),
        '',
        '',
        '',
        status,
    ])


def append_comment_rows(out_ws, note, note_id, comments, status):
    for index, comment in enumerate(comments, start=1):
        row_status = status
        if not comment.get('id') or comment.get('content') is None:
            row_status = '失败'
        out_ws.append([
            note.get('笔记链接', ''),
            index,
            build_single_comment_field(comment),
            build_replies_field(comment),
            row_status,
        ])


def collect_valid_notes(ws, headers, progress_col):
    required = ['采集状态', '笔记链接']
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(f'notes_raw 缺少必要字段：{", ".join(missing)}')

    seen_note_ids = set()
    notes = []
    for row in range(2, ws.max_row + 1):
        status = str(cell(ws, row, headers, '采集状态') or '').strip()
        link = str(cell(ws, row, headers, '笔记链接') or '').strip()
        if status != '成功' or not link:
            continue

        try:
            note_id, _ = parse_note_url(link)
        except Exception:
            note_id = ''

        dedupe_key = note_id or link
        if str(ws.cell(row, progress_col).value or '').strip() == DONE_MARK:
            seen_note_ids.add(dedupe_key)
            continue
        if dedupe_key in seen_note_ids:
            ws.cell(row, progress_col).value = DONE_MARK
            continue
        seen_note_ids.add(dedupe_key)

        notes.append({
            'row': row,
            '笔记链接': link,
            'note_id': note_id,
        })
    return notes


def main():
    parser = argparse.ArgumentParser(description='基于 notes_raw 采集一级评论并输出 comments_raw')
    parser.add_argument('excel', nargs='?', default='notes_raw.xlsx')
    parser.add_argument('-o', '--output', default='comments_raw.xlsx', help='评论输出文件')
    parser.add_argument('--max-comments', type=int, default=0, help='一级评论采集上限；0 表示不限制')
    parser.add_argument('--limit-notes', type=int, default=0, help='仅用于测试：限制处理的笔记数量')
    parser.add_argument('--sleep', type=float, default=35, help='每篇笔记采集完成后的等待秒数')
    parser.add_argument('--cookies-file', default='xhs_cookies.txt', help='登录 cookie 缓存文件，存在则不再打开浏览器获取登录态')
    args = parser.parse_args()

    excel_path = Path(args.excel)
    output_path = Path(args.output)
    backup_path = excel_path.with_name(excel_path.stem + '_备份.xlsx')
    if not backup_path.exists():
        shutil.copy2(excel_path, backup_path)

    wb = load_workbook(excel_path)
    notes_ws = wb['notes_raw'] if 'notes_raw' in wb.sheetnames else wb.active
    headers = header_map(notes_ws)
    progress_col = ensure_progress_column(notes_ws)
    notes = collect_valid_notes(notes_ws, headers, progress_col)
    if args.limit_notes:
        notes = notes[:args.limit_notes]
    out_wb, out_ws = comments_raw_sheet(output_path)
    out_wb.save(output_path)

    if not notes:
        wb.save(excel_path)
        out_wb.save(output_path)
        print('没有符合条件的笔记。', flush=True)
        return

    cookies_file = Path(args.cookies_file)
    if cookies_file.exists():
        cookies_str = cookies_file.read_text(encoding='utf-8').strip()
        print(f'已读取登录态缓存：{cookies_file}', flush=True)
    else:
        first_link = first_valid_note(notes_ws, headers) or notes[0]['笔记链接']
        print(f'获取登录态：{first_link}', flush=True)
        cookies_str = get_browser_cookies(first_link)
        cookies_file.write_text(cookies_str, encoding='utf-8')
        print(f'登录态已缓存：{cookies_file}', flush=True)

    for index, note in enumerate(notes, start=1):
        link = note['笔记链接']
        note_id = note.get('note_id') or ''
        print(f'采集 {index}/{len(notes)}：{link}', flush=True)

        try:
            parsed_note_id, xsec_token = parse_note_url(link)
            note_id = note_id or parsed_note_id
        except Exception:
            append_status_row(out_ws, note, note_id, '失败')
            notes_ws.cell(note['row'], progress_col).value = DONE_MARK
            out_wb.save(output_path)
            wb.save(excel_path)
            continue

        try:
            comments = get_all_comments_with_replies(note_id, xsec_token, cookies_str, limit=args.max_comments)
            if not comments:
                append_status_row(out_ws, note, note_id, '无评论')
            else:
                append_comment_rows(out_ws, note, note_id, comments, '成功')
            notes_ws.cell(note['row'], progress_col).value = DONE_MARK
            out_wb.save(output_path)
            wb.save(excel_path)
        except Exception as error:
            message = str(error)
            status = '加载失败' if '461' in message or 'comment' in message.lower() else '失败'
            append_status_row(out_ws, note, note_id, status)
            if status != '加载失败':
                notes_ws.cell(note['row'], progress_col).value = DONE_MARK
            out_wb.save(output_path)
            wb.save(excel_path)
            print(f'  {status}：{error}', flush=True)
            if '461' in message:
                print('检测到安全限制，已停止本轮采集。请稍后从当前行继续运行。', flush=True)
                break

        time.sleep(args.sleep)

    print(f'完成，comments_raw 已写入：{output_path}', flush=True)


if __name__ == '__main__':
    main()
